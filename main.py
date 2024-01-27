import os
import PyPDF2
import re
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl


def create_matrix(pdf_path):
    matrix = []
    kk = 1
    with open(pdf_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        num_pages = len(pdf_reader.pages)

        for page_num in range(num_pages):
            page = pdf_reader.pages[page_num]
            page_text = page.extract_text()
            lines = page_text.split('\n')
            start_index = -1
            end_index = -1

            # Find the start and end indices of the desired lines
            for i, line in enumerate(lines):
                if line.startswith('ID Color Quantity Part name'):
                    start_index = i + 1
                elif line.startswith('N'):
                    if start_index != -1:
                        end_index = i
                        break

            # Add the desired lines to the matrix without the first number
            if start_index != -1 and end_index != -1:
                matrix.append([])
                matrix[kk - 1].extend(
                    re.sub(r'^\d+\s', '', line)
                    for line in lines[start_index:end_index]
                )
                kk += 1
            print(matrix)
    arr = extract_sheet_multiplicity(pdf_path)
    matrix = update_matrix(matrix,arr)
    return matrix


def update_matrix(matrix, array):
    updated_matrix = []
    for i, row in enumerate(matrix):
        updated_row = []
        for element in row:
            num, text = element.split(" ")
            num = int(num)
            updated_num = num * array[i]
            updated_element = str(updated_num) + " " + text
            updated_row.append(updated_element)
        updated_matrix.append(updated_row)
    return updated_matrix


def extract_sheet_multiplicity(pdf_path):
    sheet_multiplicity_list = []  # Initialize an empty list

    with open(pdf_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        num_pages = len(pdf_reader.pages)

        for page_num in range(num_pages):
            page = pdf_reader.pages[page_num]
            page_text = page.extract_text()

            # Split the page text into lines
            lines = page_text.split('\n')

            # Iterate over the lines and extract the number from lines starting with "Sheet multiplicity"
            for line in lines:
                if line.startswith("Sheet multiplicity"):
                    number = re.findall(r'\d+', line)
                    if number:
                        sheet_multiplicity_list.append(int(number[0]))  # Append the extracted number as an integer

    return sheet_multiplicity_list


def create_excel_file(matrix, output_path):
    file_name = os.path.splitext(output_path)[0] + '.xlsx'

    # Create a new workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Iterate over the matrix and populate the Excel file
    row_index = 1
    for element_index, element in enumerate(matrix):
        for string in element:
            # Extract the number and text from the string
            number, text = string.split(' ', 1)
            number = int(number)

            # Write the text and element index to the Excel file
            for _ in range(number):
                sheet.cell(row=row_index, column=1).value = text
                sheet.cell(row=row_index, column=2).value = element_index + 1
                row_index += 1

    # Save the workbook to the specified output path
    workbook.save(file_name)


def browse_pdf_file():
    file_path = filedialog.askopenfilename(filetypes=[('PDF Files', '*.pdf')])
    entry_pdf_path.delete(0, tk.END)
    entry_pdf_path.insert(0, file_path)


def execute_conversion():
    pdf_path = entry_pdf_path.get()
    if pdf_path:
        try:
            matrix = create_matrix(pdf_path)
            output_path = os.path.splitext(pdf_path)[0] + '.xlsx'
            create_excel_file(matrix, output_path)
            messagebox.showinfo('Success', 'Conversion completed successfully.')
        except Exception as e:
            messagebox.showerror('Error', str(e))
    else:
        messagebox.showwarning('Warning', 'Please select a PDF file.')


# Create the main window
window = tk.Tk()
window.title('PDF to Excel Conversion')

# Create and position the file selection widgets
label_pdf_path = tk.Label(window, text='PDF File:')
label_pdf_path.grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)

entry_pdf_path = tk.Entry(window, width=50)
entry_pdf_path.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)

button_browse = tk.Button(window, text='Browse', command=browse_pdf_file)
button_browse.grid(row=0, column=2, padx=5, pady=5)

# Create and position the conversion button
button_execute = tk.Button(window, text='Execute Conversion', command=execute_conversion)
button_execute.grid(row=1, column=0, columnspan=3, padx=5, pady=5)

# Start the main event loop
window.mainloop()
import os
import PyPDF2
import re
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl


def create_matrix(pdf_path):
    matrix = []
    kk = 1
    with open(pdf_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        num_pages = len(pdf_reader.pages)

        for page_num in range(num_pages):
            page = pdf_reader.pages[page_num]
            page_text = page.extract_text()
            lines = page_text.split('\n')
            start_index = -1
            end_index = -1

            # Find the start and end indices of the desired lines
            for i, line in enumerate(lines):
                if line.startswith('ID Color Quantity Part name'):
                    start_index = i + 1
                elif line.startswith('N'):
                    if start_index != -1:
                        end_index = i
                        break

            # Add the desired lines to the matrix without the first number
            if start_index != -1 and end_index != -1:
                matrix.append([])
                matrix[kk - 1].extend(
                    re.sub(r'^\d+\s', '', line)
                    for line in lines[start_index:end_index]
                )
                kk += 1
            print(matrix)
    arr = extract_sheet_multiplicity(pdf_path)
    matrix = update_matrix(matrix,arr)
    return matrix


def update_matrix(matrix, array):
    updated_matrix = []
    for i, row in enumerate(matrix):
        updated_row = []
        for element in row:
            num, text = element.split(" ")
            num = int(num)
            updated_num = num * array[i]
            updated_element = str(updated_num) + " " + text
            updated_row.append(updated_element)
        updated_matrix.append(updated_row)
    return updated_matrix


def extract_sheet_multiplicity(pdf_path):
    sheet_multiplicity_list = []  # Initialize an empty list

    with open(pdf_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        num_pages = len(pdf_reader.pages)

        for page_num in range(num_pages):
            page = pdf_reader.pages[page_num]
            page_text = page.extract_text()

            # Split the page text into lines
            lines = page_text.split('\n')

            # Iterate over the lines and extract the number from lines starting with "Sheet multiplicity"
            for line in lines:
                if line.startswith("Sheet multiplicity"):
                    number = re.findall(r'\d+', line)
                    if number:
                        sheet_multiplicity_list.append(int(number[0]))  # Append the extracted number as an integer

    return sheet_multiplicity_list


def create_excel_file(matrix, output_path):
    file_name = os.path.splitext(output_path)[0] + '.xlsx'

    # Create a new workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Iterate over the matrix and populate the Excel file
    row_index = 1
    for element_index, element in enumerate(matrix):
        for string in element:
            # Extract the number and text from the string
            number, text = string.split(' ', 1)
            number = int(number)

            # Write the text and element index to the Excel file
            for _ in range(number):
                sheet.cell(row=row_index, column=1).value = text
                sheet.cell(row=row_index, column=2).value = element_index + 1
                row_index += 1

    # Save the workbook to the specified output path
    workbook.save(file_name)


def browse_pdf_file():
    file_path = filedialog.askopenfilename(filetypes=[('PDF Files', '*.pdf')])
    entry_pdf_path.delete(0, tk.END)
    entry_pdf_path.insert(0, file_path)


def execute_conversion():
    pdf_path = entry_pdf_path.get()
    if pdf_path:
        try:
            matrix = create_matrix(pdf_path)
            output_path = os.path.splitext(pdf_path)[0] + '.xlsx'
            create_excel_file(matrix, output_path)
            messagebox.showinfo('Success', 'Conversion completed successfully.')
        except Exception as e:
            messagebox.showerror('Error', str(e))
    else:
        messagebox.showwarning('Warning', 'Please select a PDF file.')


# Create the main window
window = tk.Tk()
window.title('PDF to Excel Conversion')

# Create and position the file selection widgets
label_pdf_path = tk.Label(window, text='PDF File:')
label_pdf_path.grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)

entry_pdf_path = tk.Entry(window, width=50)
entry_pdf_path.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)

button_browse = tk.Button(window, text='Browse', command=browse_pdf_file)
button_browse.grid(row=0, column=2, padx=5, pady=5)

# Create and position the conversion button
button_execute = tk.Button(window, text='Execute Conversion', command=execute_conversion)
button_execute.grid(row=1, column=0, columnspan=3, padx=5, pady=5)

# Start the main event loop
window.mainloop()
